import os
import json
import openai
import pandas as pd
from dotenv import load_dotenv
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import glob
import threading
import time
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from utils import send_to_api
import email.utils

load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")
client = openai

def parse_twitter_date(date_string):
    """Parse Twitter-style date format and convert to ISO format"""
    if not date_string:
        return ""
    
    try:
        # Parse the Twitter-style date format: "Mon Jul 28 17:12:07 +0000 2025"
        parsed_date = email.utils.parsedate_to_datetime(date_string)
        # Convert to ISO format
        return parsed_date.isoformat()
    except Exception as e:
        print(f"Warning: Could not parse date '{date_string}': {e}")
        # Return original string if parsing fails
        return date_string

def get_fire_analysis(content):
    """Get fire score, state, and county from tweet content"""
    prompt = (
        "Analyze the following tweet about fire incidents in the United States and provide:\n"
        "1. A fire-relatedness score from 0-10 (0=not related, 10=definitely fire-related)\n"
        "2. The US state where the fire occurred (use ONLY the state name like California, Texas, New York, Arizona, etc. - do NOT include 'County' in the state name)\n"
        "3. The county where the fire occurred (use ONLY the county name, not the state name)\n\n"
        "Respond in this exact format:\n"
        "Score: [0-10]\n"
        "State: [State name only or N/A]\n"
        "County: [County name only or N/A]\n\n"
        "Examples:\n"
        "- If tweet mentions 'Arizona', State should be 'Arizona', County should be the specific county or N/A\n"
        "- If tweet mentions 'Los Angeles County, California', State should be 'California', County should be 'Los Angeles'\n"
        "- If location cannot be determined, use N/A for both state and county.\n\n"
        f"Tweet content: {content[:2000]}"
    )
    messages = [
        {"role": "system", "content": "You are an AI that analyzes fire-related tweets to extract fire scores and location information. Always respond in the exact format specified."},
        {"role": "user", "content": prompt}
    ]
    try:
        ai_response = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=messages,
            temperature=0,
        )
        answer = ai_response.choices[0].message.content.strip()
        
        # Parse the response
        score_match = re.search(r'Score:\s*(\d+)', answer)
        state_match = re.search(r'State:\s*([A-Za-z\s]+|N/A)', answer)
        county_match = re.search(r'County:\s*([^\n]+)', answer)
        
        score = int(score_match.group(1)) if score_match else 0
        state = state_match.group(1).strip() if state_match else "N/A"
        county = county_match.group(1).strip() if county_match else "N/A"
        
        # Clean up state and county - remove any extra whitespace or newlines
        state = re.sub(r'\s+', ' ', state).strip()
        county = re.sub(r'\s+', ' ', county).strip()
        
        # Fix common parsing errors
        # If state contains "County", extract just the state name
        if 'County' in state and state != 'N/A':
            # Try to extract state name before "County"
            state_parts = state.split('County')
            if state_parts[0].strip():
                state = state_parts[0].strip()
        
        # If county is "N/A" but state contains county info, try to extract it
        if county == 'N/A' and 'County' in state and state != 'N/A':
            # This might be a case where the AI put county info in state field
            county_parts = state.split('County')
            if len(county_parts) > 1 and county_parts[1].strip():
                county = county_parts[1].strip()
            elif len(county_parts) > 0:
                county = county_parts[0].strip()
        
        # Validate state names - ensure they're actual US states
        valid_states = [
            'Alabama', 'Alaska', 'Arizona', 'Arkansas', 'California', 'Colorado', 'Connecticut',
            'Delaware', 'Florida', 'Georgia', 'Hawaii', 'Idaho', 'Illinois', 'Indiana', 'Iowa',
            'Kansas', 'Kentucky', 'Louisiana', 'Maine', 'Maryland', 'Massachusetts', 'Michigan',
            'Minnesota', 'Mississippi', 'Missouri', 'Montana', 'Nebraska', 'Nevada', 'New Hampshire',
            'New Jersey', 'New Mexico', 'New York', 'North Carolina', 'North Dakota', 'Ohio',
            'Oklahoma', 'Oregon', 'Pennsylvania', 'Rhode Island', 'South Carolina', 'South Dakota',
            'Tennessee', 'Texas', 'Utah', 'Vermont', 'Virginia', 'Washington', 'West Virginia',
            'Wisconsin', 'Wyoming'
        ]
        
        # If state doesn't match a valid state name, set to N/A
        if state not in valid_states and state != 'N/A':
            print(f"[WARNING] Invalid state name detected: '{state}', setting to N/A")
            state = 'N/A'
        
        return score, state, county
    except Exception as e:
        print(f"Error with OpenAI API (analysis): {e}")
        return 0, "N/A", "N/A"

def get_fire_related_score(content):
    """Get a score from 0-10 indicating how fire-related the tweet is (legacy function)"""
    score, _, _ = get_fire_analysis(content)
    return score

def verify_fire_incident(text, url):
    """Verify if the tweet describes a fire incident in the USA"""
    print(f"Verifying: {url}")
    truncated_content = text[:4000]
    fire_incident_prompt = (
        "You are given the content of a tweet. Determine if it describes a fire incident in the United States that likely caused damage to physical structures (such as homes, apartments, offices, commercial buildings, factories, or infrastructure). "
        "The fire may have resulted in structural damage or destruction, due to causes like electrical faults, negligence, accidents, natural disasters (e.g., wildfires), or arson. "
        "Be inclusive: If the tweet suggests a fire incident with possible or likely damage to structures, even if not 100% explicit, respond with 'yes'. "
        "Respond with 'yes' if the tweet is about a fire incident in the USA that could have caused damage to physical structures. Otherwise, respond with 'no'.\n\n"
        f"Tweet content: {truncated_content}\nURL: {url}\n"
        "Only use the provided content for your evaluation. Do not infer or assume details not present in the text, but err on the side of inclusion if the fire incident is plausible."
    )
    messages = [
        {
            "role": "system",
            "content": "You are an AI tasked with evaluating tweets to determine if they describe fire damages or destruction in the United States. Be inclusive: If the tweet is plausibly about fire damages or destruction in the USA, mark as 'yes'."
        },
        {"role": "user", "content": fire_incident_prompt}
    ]
    try:
        ai_response = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=messages,
            temperature=0,
        )
        answer = ai_response.choices[0].message.content.strip()
        print(f"Result: {answer}")
        return answer
    except Exception as e:
        print(f"Error with OpenAI API: {e}")
        return "no"

def update_live_json(live_json_path, entry):
    """Thread-safe function to update the live JSON file"""
    lock = threading.Lock()
    with lock:
        try:
            if os.path.exists(live_json_path):
                with open(live_json_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            else:
                data = []
            
            # Check if entry already exists (by tweet ID)
            existing_ids = [item.get('tweet_id') for item in data]
            tweet_id = entry.get('tweet_id')
            
            if tweet_id not in existing_ids:
                data.append(entry)
                
                with open(live_json_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                print(f"[OK] Live JSON updated: {tweet_id} (total items: {len(data)})")
            else:
                print(f"[SKIP] Tweet {tweet_id} already exists in JSON (total items: {len(data)})")
        except Exception as e:
            print(f"Error updating live JSON: {e}")

def update_excel_file(excel_path, new_row):
    """Update Excel file with new verified tweet"""
    try:
        if os.path.exists(excel_path):
            # Read existing data
            df_existing = pd.read_excel(excel_path)
            # Add new row
            df_new = pd.DataFrame([new_row])
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        else:
            # Create new DataFrame
            df_combined = pd.DataFrame([new_row])
        
        # Save to Excel
        df_combined.to_excel(excel_path, index=False)
        
        # Format Excel file
        autosize_and_format_excel(excel_path)
        print(f"[EXCEL] Excel updated: {new_row.get('tweet_id')}")
        
    except Exception as e:
        print(f"Error updating Excel: {e}")

def autosize_and_format_excel(excel_path):
    """Format Excel file with proper column widths and hyperlinks"""
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Set column widths
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_length = 0
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(15, min(60, max_length + 2))
            for cell in col:
                cell.alignment = cell.alignment.copy(wrap_text=True)
        
        # Set row heights
        for row in ws.iter_rows():
            max_height = 15
            for cell in row:
                if cell.value:
                    lines = str(cell.value).count("\n") + 1
                    length = len(str(cell.value))
                    est_height = max(15, min(150, lines * 15 + length // 50 * 15))
                    if est_height > max_height:
                        max_height = est_height
            ws.row_dimensions[row[0].row].height = max_height
        
        # Add hyperlinks to URL column
        url_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value and str(cell.value).lower() == "url":
                url_col = idx
                break
        
        if url_col:
            for row in ws.iter_rows(min_row=2, min_col=url_col, max_col=url_col):
                for cell in row:
                    if cell.value and str(cell.value).startswith("http"):
                        cell.hyperlink = cell.value
                        cell.style = "Hyperlink"
        
        wb.save(excel_path)
        
    except Exception as e:
        print(f"Error formatting Excel: {e}")

def send_email_results(excel_path, json_path, verified_count):
    """Send verification results via email"""
    try:
        # Email configuration
        sender_email = os.getenv("EMAIL_ADDRESS")
        sender_password = os.getenv("EMAIL_PASSWORD")
        smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        
        if not sender_email or not sender_password:
            print("[ERROR] Email credentials not found in environment variables!")
            print("Please set EMAIL_ADDRESS and EMAIL_PASSWORD in your .env file")
            return
        
        # Recipient emails
        recipient_emails = [
            "info@theagilemorph.com",
            "forrohitsingh99@gmail.com",
            "unipaney@dhaninfo.biz",
            "u@agilemorph.biz", 
            "rchakraborty@dhaninfo.biz",
            "npalliwal@dhaninfo.biz",
            "lalit.shukla@dhaninfo.biz",
            "rnagmote@dhaninfo.biz",
            "apandey@dhaninfo.biz",
            "careports@dhaninfo.biz"
        ]
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ", ".join(recipient_emails)
        msg['Subject'] = f"Fire Incident Verification Results - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Email body with HTML formatting and AgileMorph branding
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    max-width: 600px;
                    margin: 0 auto;
                    padding: 20px;
                    background-color: #f9f9f9;
                }}
                .header {{
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    padding: 30px;
                    text-align: center;
                    border-radius: 10px 10px 0 0;
                    margin-bottom: 0;
                }}
                .header h1 {{
                    margin: 0;
                    font-size: 28px;
                    font-weight: 600;
                }}
                .content {{
                    background: white;
                    padding: 30px;
                    border-radius: 0 0 10px 10px;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                }}
                .summary-box {{
                    background: #f8f9fa;
                    border-left: 4px solid #667eea;
                    padding: 20px;
                    margin: 20px 0;
                    border-radius: 5px;
                }}
                .summary-box h3 {{
                    margin-top: 0;
                    color: #667eea;
                    font-size: 18px;
                }}
                .stats {{
                    display: flex;
                    justify-content: space-between;
                    margin: 20px 0;
                }}
                .stat-item {{
                    text-align: center;
                    flex: 1;
                    padding: 15px;
                    background: #f8f9fa;
                    margin: 0 5px;
                    border-radius: 5px;
                }}
                .stat-number {{
                    font-size: 24px;
                    font-weight: bold;
                    color: #667eea;
                }}
                .stat-label {{
                    font-size: 12px;
                    color: #666;
                    text-transform: uppercase;
                }}
                .footer {{
                    text-align: center;
                    margin-top: 30px;
                    padding-top: 20px;
                    border-top: 1px solid #eee;
                }}
                .brand-link {{
                    display: inline-block;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    text-decoration: none;
                    padding: 12px 25px;
                    border-radius: 25px;
                    font-weight: 600;
                    margin: 10px 0;
                    transition: transform 0.2s;
                }}
                .brand-link:hover {{
                    transform: translateY(-2px);
                }}
                .attachments {{
                    background: #e8f4fd;
                    border: 1px solid #bee5eb;
                    border-radius: 5px;
                    padding: 15px;
                    margin: 20px 0;
                }}
                .attachments h4 {{
                    margin-top: 0;
                    color: #0c5460;
                }}
                .attachments ul {{
                    margin: 10px 0;
                    padding-left: 20px;
                }}
                .attachments li {{
                    margin: 5px 0;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>🔥 Fire Incident Detection Report</h1>
                <p>Automated AI-Powered Fire Incident Verification</p>
            </div>
            
            <div class="content">
                <div class="summary-box">
                    <h3>📊 Executive Summary</h3>
                    <p>Your automated fire incident detection system has completed its analysis of social media data from the last 72 hours.</p>
                </div>
                
                <div class="stats">
                    <div class="stat-item">
                        <div class="stat-number">{verified_count}</div>
                        <div class="stat-label">Verified Incidents</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-number">{datetime.now().strftime('%H:%M')}</div>
                        <div class="stat-label">Completion Time</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-number">{datetime.now().strftime('%m/%d')}</div>
                        <div class="stat-label">Report Date</div>
                    </div>
                </div>
                
                <div class="attachments">
                    <h4>📎 Attached Files</h4>
                    <ul>
                        <li><strong>Excel Report:</strong> Detailed analysis with verified fire incidents</li>
                        <li><strong>JSON Data:</strong> Raw data for further processing</li>
                    </ul>
                </div>
                
                <p><strong>Report Coverage:</strong> This automated analysis covers fire-related social media activity across all 50 U.S. states, including verified incidents with structural damage potential.</p>
                
                <div class="footer">
                    <p><em>Powered by advanced AI automation and real-time social media monitoring</em></p>
                    <a href="https://theagilemorph.com/" class="brand-link" target="_blank">
                        🚀 Powered by AgileMorph
                    </a>
                    <p style="font-size: 12px; color: #666; margin-top: 15px;">
                        <strong>AgileMorph</strong> - Let AI seamlessly elevate your brand<br>
                        AI Automation • Web Development • Digital Marketing
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Create both HTML and plain text versions
        msg.attach(MIMEText(html_body, 'html'))
        
        # Plain text fallback
        plain_body = f"""
Fire Incident Verification Complete!

Summary:
- Total verified fire incidents: {verified_count}
- Verification completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Files attached:
1. Excel file with detailed results
2. JSON file with raw data

This automated report contains verified fire-related tweets from the last 72 hours.

Powered by AgileMorph - https://theagilemorph.com/
        """
        
        msg.attach(MIMEText(plain_body, 'plain'))
        
        # Attach Excel file
        if os.path.exists(excel_path):
            with open(excel_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {os.path.basename(excel_path)}'
            )
            msg.attach(part)
        
        # Attach JSON file
        if os.path.exists(json_path):
            with open(json_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {os.path.basename(json_path)}'
            )
            msg.attach(part)
        
        # Send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        text = msg.as_string()
        server.sendmail(sender_email, recipient_emails, text)
        server.quit()
        
        print(f"[EMAIL] Email sent successfully to {len(recipient_emails)} recipients!")
        
        # Send data to API endpoint
        print(f"\n[API] Sending data to API endpoint...")
        api_success = send_to_api(json_path, verified_count)
        if api_success:
            print(f"[API] Data successfully sent to API endpoint")
        else:
            print(f"[API] Failed to send data to API endpoint")
        
    except Exception as e:
        print(f"[ERROR] Error sending email: {e}")
        print("Please check your email configuration in .env file")

def verify_and_save_tweets(cleaned_json_path, output_dir="output"):
    """Main function to verify tweets and save results live"""
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate timestamped filenames with more detail
    dt_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = os.path.join(output_dir, f"verified_fires_{dt_str}.xlsx")
    live_json_path = os.path.join(output_dir, f"live_verified_fires_{dt_str}.json")
    
    print(f"[OUTPUT] Output files:")
    print(f"   Excel: {excel_path}")
    print(f"   JSON: {live_json_path}")
    
    # Load cleaned tweets
    try:
        with open(cleaned_json_path, "r", encoding="utf-8") as f:
            tweets = json.load(f)
        print(f"[DATA] Loaded {len(tweets)} tweets for verification")
    except Exception as e:
        print(f"Error loading tweets: {e}")
        return
    
    verified_count = 0
    
    # Process each tweet
    for i, tweet in enumerate(tqdm(tweets, desc="Verifying tweets with AI")):
        try:
            # Extract tweet data
            tweet_id = tweet.get('id', f"tweet_{i}")
            text = tweet.get('text', '')
            url = tweet.get('url', '')
            created_at_raw = tweet.get('createdAt', '')
            # Parse and format the date properly
            created_at = parse_twitter_date(created_at_raw)
            author = tweet.get('author', {})
            username = author.get('userName', 'Unknown') if author else 'Unknown'
            
            # Skip if no text or too short
            if not text.strip():
                continue
            
            # Skip tweets with content less than 20 characters
            if len(text.strip()) < 20:
                print(f"[SKIP] Tweet {tweet_id} too short ({len(text.strip())} chars): {text[:50]}...")
                continue
            
            print(f"[DEBUG] Processing tweet {tweet_id} - Length: {len(text.strip())} chars")
            
            # Verify with AI
            verification_result = verify_fire_incident(text, url)
            
            # If verified, get fire analysis (score, state, county) and save
            if verification_result.lower().startswith("yes"):
                fire_score, state, county = get_fire_analysis(text)
                print(f"[DEBUG] Tweet {tweet_id} - Fire score: {fire_score}, State: '{state}', County: '{county}'")
                
                # Only save tweets with fire score >= 5
                if fire_score >= 5:
                    verified_at = datetime.now().isoformat()
                    
                    # Create entry with all necessary columns including tweet_id for deduplication
                    entry = {
                        'tweet_id': tweet_id,
                        'title': text[:100] + "..." if len(text) > 100 else text,
                        'content': text,
                        'published_date': created_at,
                        'url': url,
                        'source': username,
                        'fire_related_score': fire_score,
                        'state': state,
                        'county': county,
                        'verification_result': verification_result,
                        'verified_at': verified_at
                    }
                    
                    # Save to live JSON immediately
                    update_live_json(live_json_path, entry)
                    
                    # Update Excel file
                    update_excel_file(excel_path, entry)
                    
                    verified_count += 1
                    print(f"[FIRE] Verified tweet {verified_count} (score: {fire_score}): {tweet_id}")
                    
                    # Small delay to show live processing
                    time.sleep(0.5)
                else:
                    print(f"[SKIP] Tweet {tweet_id} verified but score too low ({fire_score}/10) - skipping")
            
        except Exception as e:
            print(f"Error processing tweet {i}: {e}")
            continue
    
    print(f"\n[SUCCESS] Verification complete!")
    print(f"[OK] Total verified fire incidents: {verified_count}")
    print(f"[OUTPUT] Results saved to:")
    print(f"   Excel: {excel_path}")
    print(f"   JSON: {live_json_path}")
    
    return verified_count, excel_path, live_json_path

def fix_existing_json_dates(json_file_path):
    """Fix date formats in existing JSON files"""
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        fixed_count = 0
        for item in data:
            if 'published_date' in item and item['published_date']:
                original_date = item['published_date']
                fixed_date = parse_twitter_date(original_date)
                if fixed_date != original_date:
                    item['published_date'] = fixed_date
                    fixed_count += 1
        
        if fixed_count > 0:
            # Save the fixed data back to the file
            with open(json_file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"[FIXED] Updated {fixed_count} date entries in {json_file_path}")
        else:
            print(f"[INFO] No date fixes needed in {json_file_path}")
            
    except Exception as e:
        print(f"Error fixing dates in {json_file_path}: {e}")

def main():
    """Main execution function"""
    import sys
    
    # Determine input file
    if len(sys.argv) > 1:
        json_path = sys.argv[1]
    else:
        # Look for most recent fire_tweets_72h_*.json file first, then cleaned files
        fire_tweets_files = glob.glob("fire_tweets_72h_*.json")
        if fire_tweets_files:
            # Use the most recent one
            json_path = max(fire_tweets_files, key=os.path.getctime)
            print(f"[FILE] Using latest fire tweets file: {json_path}")
        else:
            # Fallback to cleaned tweet files
            cleaned_files = glob.glob("*cleaned*.json")
            if cleaned_files:
                # Use the most recent one
                json_path = max(cleaned_files, key=os.path.getctime)
                print(f"[FILE] Using latest cleaned file: {json_path}")
            else:
                print("[ERROR] No fire_tweets_72h_*.json or cleaned tweets file found!")
                print("Please run tweet_fire_search.py first or specify a file path.")
                print("Usage: python verify_tweets.py [path_to_tweets.json]")
                return
    
    if not os.path.exists(json_path):
        print(f"[ERROR] File not found: {json_path}")
        return
    
    # Check for OpenAI API key
    if not os.getenv("OPENAI_API_KEY"):
        print("[ERROR] OPENAI_API_KEY not found in environment variables!")
        print("Please set your OpenAI API key in a .env file or environment variable.")
        return
    
    print(f"[START] Starting tweet verification process...")
    print(f"[FILE] Input file: {json_path}")
    
    # Fix date formats in the input file if needed
    print(f"[DATE] Checking and fixing date formats...")
    fix_existing_json_dates(json_path)
    
    # Run verification
    verified_count, excel_path, json_path = verify_and_save_tweets(json_path)
    
    if verified_count > 0:
        print(f"\n[EMAIL] Sending results via email...")
        send_email_results(excel_path, json_path, verified_count)
    else:
        print(f"\n[EMAIL] No verified incidents found - no email sent.")

if __name__ == "__main__":
    main() 